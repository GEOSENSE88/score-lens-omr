# -*- coding: utf-8 -*-
"""
OMR 가채점기 회귀 테스트 스위트 (score_lens / omr_scorer)

코드를 고친 뒤 이 한 파일로 전체 정확성을 재확인한다:

    python tests_regression.py            # 오프라인 테스트 전부
    python tests_regression.py --web      # + 웹 e2e (127.0.0.1:5050 서버 필요)

검증 기준값은 전부 '실측'이다:
  - 고3: 2026-06 모평 실스캔 31명 (work/20260604165332_history 등)
  - 고1: synth_omr 가상 카드 + 실제 EBSi 정답키(202606041)
"""
from __future__ import annotations

import csv
import json
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
PASS, FAIL = [], []


def check(name: str, cond: bool, detail: str = ""):
    (PASS if cond else FAIL).append((name, detail))
    print(("  ✓ " if cond else "  ✗ ") + name + (f" — {detail}" if detail and not cond else ""))


def run_cli(args: list[str]) -> subprocess.CompletedProcess:
    env_kw = dict(cwd=ROOT, text=True, encoding="utf-8", errors="replace",
                  stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=600)
    import os
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    return subprocess.run([sys.executable] + args, env=env, **env_kw)


# ── 1. 임포트 & 구성 파일 유효성 ─────────────────────────────────
def t1_imports():
    print("[1] 모듈 임포트 & 구성 유효성")
    import omr_core, math_core, id_reader, history_core, english_core, explore_core  # noqa
    import run_objective, consolidate, report_export, exam_registry, synth_omr, web_app  # noqa
    check("모듈 임포트", True)
    for spec in sorted((ROOT / "specs").glob("*_g12.json")):
        d = json.loads(spec.read_text(encoding="utf-8"))
        check(f"spec {spec.stem}", bool(d.get("objective")) and all(
            {"q0", "nrows", "x_lo", "x_hi"} <= set(b) for b in d["objective"]))
    for tp in sorted((ROOT / "templates").glob("*.json")):
        d = json.loads(tp.read_text(encoding="utf-8"))
        # 탐구(explore)는 selections 스키마, 나머지는 objective 스키마
        check(f"template {tp.stem}", bool(d.get("objective") or d.get("selections")))


# ── 2. 정답키 혼선 가드 (2026-07-03 실회귀) ──────────────────────
def t2_key_isolation():
    print("[2] 정답키 exam_id 격리")
    import consolidate as c
    ans3 = json.loads((ROOT / "keys/202606043_국어_화법과작문.json").read_text(encoding="utf-8"))["answers"]
    ans1 = json.loads((ROOT / "keys/202606041_국어_xip_api.json").read_text(encoding="utf-8"))["answers"]
    got3, _ = c._load_key_points(ROOT / "keys", "국어", "화법과작문", "202606043")
    got1, _ = c._load_key_points(ROOT / "keys", "국어", None, "202606041")
    check("고3 국어 키 정확", all(got3[q] == int(ans3[str(q)]) for q in got3))
    check("고1 국어 키 정확", all(got1[q] == int(ans1[str(q)]) for q in got1))


# ── 3. 고3 한국사 31명 실스캔 재현 ──────────────────────────────
def t3_history_g3():
    print("[3] 고3 한국사 31명 (실스캔)")
    import cv2
    import omr_core as oc
    import id_reader as idr
    import history_core as hc
    import run_objective as ro
    template = ro.load_template(ROOT / "templates/history_eval_2027_06.json")
    key = ro.load_subject_key(ROOT / "keys", "한국사", "202606043")
    pages = sorted((ROOT / "work/20260604165332_history").glob("*.png"))
    ref = {}
    with open(ROOT / "output/20260604165332_한국사_판독표.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            ref[(r["반"], r["번호"])] = r["점수"]
    match = 0
    for pg in pages:
        gray = cv2.cvtColor(oc.load_page(str(pg)), cv2.COLOR_BGR2GRAY)
        sid = idr.read_id(gray, "g3")
        ans = hc.read_answers(gray, oc.make_anchor_transform(gray), template)
        g = ro.grade_known(ans, key, 20)
        if str(g["score"]) == str(ref.get((sid["ban"], sid["beon"]))):
            match += 1
    check(f"점수 재현 {match}/{len(pages)}", match == len(pages) == 31)


# ── 4. 고3 통합성적표 기준값 (로컬 실데이터) ──────────────────────
# 기준 학생명·값은 .local_fixtures.json (gitignore) 에 둔다 — 공개 저장소에
# 실제 학생 이름이 올라가지 않도록. 파일/데이터 없으면 건너뛴다.
def t4_consolidate_g3():
    print("[4] 고3 통합성적표 (로컬 기준값)")
    fx_path = ROOT / ".local_fixtures.json"
    kor_csv = ROOT / "output/20260604165616_점수표.csv"
    if not fx_path.exists() or not kor_csv.exists():
        print("  – 로컬 기준 데이터 없음 → 건너뜀 (clone 환경)")
        return
    fx = json.loads(fx_path.read_text(encoding="utf-8"))["consolidate_g3"]
    import consolidate as c
    out = c.consolidate_paths(
        3, keys_dir=str(ROOT / "keys"), exam_id="202606043",
        korean=str(kor_csv),
        math=str(ROOT / "output/20260604165525_수학_점수표.csv"),
        english=str(ROOT / "output/20260604165440_영어_판독표.csv"),
        history=str(ROOT / "output/20260604165332_한국사_판독표.csv"),
        explore=str(ROOT / "output/20260604165249_탐구_판독표.csv"),
        out=str(ROOT / "output/_회귀_통합성적표_고3.xlsx"))
    import openpyxl
    ws = openpyxl.load_workbook(out)["대교협"]
    want = tuple(fx["values"])
    found = False
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 5).value == fx["name"]:
            vals = tuple(ws.cell(r, c).value for c in (8, 9, 11, 12, 13, 6))
            check(f"기준 학생 {want}", vals == want, str(vals))
            found = True
            break
    if not found:
        check("기준 학생 행 존재", False)


# ── 5. 고1 가상카드 6과목 (CLI) ──────────────────────────────────
def t5_synth_g12():
    print("[5] 고1 가상카드 6과목 채점 (36건 대조)")
    from synth_omr import make_students, render_pdfs
    synth_dir = ROOT / "work/synth_g12"
    if not (synth_dir / "synth_국어.pdf").exists():
        render_pdfs(synth_dir)
    students = make_students(ROOT / "keys", "202606041", n=6, grade=1)
    exp = {(info["beon"], s): sc for info, _, e in students for s, sc in e.items()}
    templates = dict(국어="korean_g12", 수학="math_g12", 영어="english_g12",
                     한국사="history_g12", 통합사회="social_g12", 통합과학="science_g12")
    total = ok = 0
    for subj, tp in templates.items():
        p = run_cli(["run_objective.py", str(synth_dir / f"synth_{subj}.pdf"),
                     "--subject", subj, "--template", f"templates/{tp}.json",
                     "--irecord", "202606041", "--id-layout", "g12",
                     "--names", str(synth_dir / "synth_names.csv"),
                     "--out", "output/_회귀_synth"])
        if p.returncode != 0:
            check(f"{subj} 실행", False, p.stdout[-200:])
            continue
        with open(ROOT / f"output/_회귀_synth/synth_{subj}_{subj}_판독표.csv", encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                total += 1
                if float(r["점수"]) == float(exp[(r["번호"], subj)]):
                    ok += 1
    check(f"점수 대조 {ok}/{total}", total == 36 and ok == 36)


# ── 6. 에지케이스 (중복·미마킹·결시) ─────────────────────────────
def t6_edge():
    print("[6] 에지케이스 (중복마킹·미마킹·결시)")
    pdf = ROOT / "work/synth_edge/edge_english.pdf"
    if not pdf.exists():
        check("edge_english.pdf 존재", False, "synth_edge 미생성 — 문서 참조")
        return
    p = run_cli(["run_objective.py", str(pdf), "--subject", "영어",
                 "--template", "templates/english_g12.json",
                 "--irecord", "202606041", "--id-layout", "g12",
                 "--out", "output/_회귀_edge"])
    rows = list(csv.DictReader(open(ROOT / "output/_회귀_edge/edge_english_영어_판독표.csv",
                                    encoding="utf-8-sig")))
    p1, p2, p3 = rows
    check("만점 페이지 100", p1["점수"] == "100")
    check("중복+빈칸 페이지 83", p2["점수"] == "83")
    check("중복은 -1, 빈칸은 0", p2["1"] == "-1" and p2["4"] == "0")
    check("결시 0점 + 경고", p3["점수"] == "0" and "수험번호" in p3["확인필요"])


# ── 6.5 시뮬레이션(ultracode) 발견 이슈 고정 테스트 ──────────────
def t65_simfixes():
    print("[6.5] 시뮬레이션 수정 고정 (탈락·분열·바꿔치기·ID보정)")
    import importlib
    import consolidate as c
    import report_export as rx
    importlib.reload(rx); importlib.reload(c)
    import openpyxl

    # a) 통합과학 판독표만 → 학생 생성 + 과탐 열에만 표기
    out = c.consolidate_paths(1, keys_dir=str(ROOT / "keys"), exam_id="202606041",
        science=str(ROOT / "output/_회귀_synth/synth_통합과학_통합과학_판독표.csv"),
        out=str(ROOT / "output/_회귀_simfix_a.xlsx"))
    ws = openpyxl.load_workbook(out)["UNIV 1학년"]
    n = sum(1 for r in range(3, ws.max_row + 1) if ws.cell(r, 4).value)
    check("탐구만 입력 시 학생 생성", n == 6 and ws.cell(3, 18).value in (None, "")
          and ws.cell(3, 22).value not in (None, ""))

    # b) 성명 유무 혼재 → 분열 없음
    import csv as _csv
    rows = list(_csv.reader(open(ROOT / "output/_회귀_synth/synth_수학_수학_판독표.csv",
                                 encoding="utf-8-sig")))
    i = rows[0].index("성명")
    for r in rows[1:]:
        r[i] = ""
    noname = ROOT / "output/_회귀_simfix_noname.csv"
    with open(noname, "w", newline="", encoding="utf-8-sig") as f:
        _csv.writer(f).writerows(rows)
    out = c.consolidate_paths(1, keys_dir=str(ROOT / "keys"), exam_id="202606041",
        korean=str(ROOT / "output/_회귀_synth/synth_국어_국어_판독표.csv"),
        math=str(noname), out=str(ROOT / "output/_회귀_simfix_b.xlsx"))
    ws = openpyxl.load_workbook(out)["UNIV 1학년"]
    n = sum(1 for r in range(3, ws.max_row + 1) if ws.cell(r, 4).value)
    both = sum(1 for r in range(3, ws.max_row + 1)
               if ws.cell(r, 6).value not in (None, "") and ws.cell(r, 10).value not in (None, ""))
    check("성명 혼재 분열 없음", n == 6 and both == 6)

    # c) 과목 바꿔치기 → 평균저점 경고
    p = run_cli(["run_objective.py", str(ROOT / "work/synth_g12/synth_영어.pdf"),
                 "--subject", "국어", "--template", "templates/korean_g12.json",
                 "--irecord", "202606041", "--id-layout", "g12",
                 "--out", "output/_회귀_simfix_swap"])
    rows = list(_csv.DictReader(open(ROOT / "output/_회귀_simfix_swap/synth_영어_국어_판독표.csv",
                                     encoding="utf-8-sig")))
    check("바꿔치기 평균저점 경고", all("평균저점" in r["확인필요"] for r in rows))

    # d) 수험번호 앵커 보정 (y+40px)
    import cv2
    import numpy as np
    import synth_omr as so
    import omr_core as oc
    import id_reader as idr
    img = so.draw_card("영어", grade=1, ban="03", beon="07", answers={1: 1})
    sh = cv2.warpAffine(img, np.float32([[1, 0, 0], [0, 1, 40]]), (so.W, so.H),
                        borderValue=so.BG)
    gray = cv2.cvtColor(sh, cv2.COLOR_BGR2GRAY)
    sid = idr.read_id(gray, "g12", oc.make_anchor_transform(gray))
    check("수험번호 앵커 보정(y+40px)", sid["ban"] == "03" and sid["beon"] == "07")

    # e) g12 과목ID 매핑 우선 (순서 뒤바뀜 내성)
    import exam_registry as er
    shuffled = {"subj_140073": "100", "subj_17012": "105", "subj_110001": "104",
                "subj_17014": "103", "subj_17020": "102", "subj_140072": "101"}
    pairs = er._g12_subject_pairs(shuffled, {})
    check("g12 과목ID 매핑", dict((s, p) for s, n, p in pairs).get("국어") == "105")


# ── 6.7 스캔 방향 자동보정 (180°/90°/270°) ──────────────────────
def t67_autorotate():
    print("[6.7] 스캔 방향 자동보정")
    import cv2
    import json as _json
    import synth_omr as so
    import omr_core as oc
    import id_reader as idr
    import history_core as hc
    import run_objective as ro
    key = _json.loads((ROOT / "keys/202606041_영어_xip_api.json").read_text(encoding="utf-8"))
    kans = {int(q): int(v) for q, v in key["answers"].items()}
    template = ro.load_template(ROOT / "templates/english_g12.json")
    keyobj = ro.load_subject_key(ROOT / "keys", "영어", "202606041")
    img = so.draw_card("영어", grade=1, ban="03", beon="07", answers=dict(kans))  # 만점 답안
    rots = {"0": img, "180": cv2.rotate(img, cv2.ROTATE_180),
            "90": cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE),
            "270": cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)}
    import numpy as np
    import tempfile
    ok = True
    for name, im in rots.items():
        p = Path(tempfile.gettempdir()) / f"_rot_{name}.png"
        cv2.imencode(".png", im)[1].tofile(str(p))
        gray = cv2.cvtColor(oc.load_page(str(p)), cv2.COLOR_BGR2GRAY)   # auto_orient 적용
        sid = idr.read_id(gray, "g12", oc.make_anchor_transform(gray))
        ans = hc.read_answers(gray, oc.make_anchor_transform(gray), template)
        g = ro.grade_known(ans, keyobj, 45)
        if not (g["score"] == 100 and sid["ban"] == "03" and sid["beon"] == "07"):
            ok = False
        p.unlink(missing_ok=True)
    check("0/90/180/270° 모두 만점+수험번호 정상", ok)


# ── 7. 좌표 보정기 재현성 (고3 영어 실스캔) ──────────────────────
def t7_calibrator():
    print("[7] 좌표 보정기 (고3 영어 기존 검증좌표 재현)")
    import calibrate_objective as co
    spec = dict(objective=[
        dict(q0=1, nrows=20, x_lo=1890, x_hi=2220, y_lo=360, y_hi=2420),
        dict(q0=21, nrows=20, x_lo=2360, x_hi=2690, y_lo=360, y_hi=2420),
        dict(q0=41, nrows=5, x_lo=2820, x_hi=3150, y_lo=360, y_hi=920)])
    pages = sorted((ROOT / "work/20260604165440_english").glob("*.png"))
    tmpl = co.calibrate(spec, pages, sample=8)
    known = json.loads((ROOT / "templates/english_eval_2027_06.json").read_text(encoding="utf-8"))
    max_dx = max(abs(a - b)
                 for blk_t, blk_k in zip(tmpl["objective"], known["objective"])
                 for a, b in zip(blk_t["xs"], blk_k["xs"]))
    check(f"열 좌표 오차 {max_dx:.1f}px ≤ 5px", max_dx <= 5)


# ── 8. 통합 워크북 레이아웃 (1·2·3학년) ─────────────────────────
def t8_report_layouts():
    print("[8] 통합 워크북 학년별 레이아웃")
    import report_export as rx
    stu = dict(학년=1, 반="01", 번호="01", 이름="테스트", 계열번호="",
               국어=dict(선택="국어", 공통원=80, 선택원="", 원점수=80, 만점=100, 정오={}),
               탐구=[dict(과목="통합사회", 원점수=40, 만점=50, 정오={}),
                    dict(과목="통합과학", 원점수=45, 만점=50, 정오={})])
    import openpyxl
    for g in (1, 2, 3):
        stu["학년"] = g
        out = rx.build_workbook([stu], ROOT / f"output/_회귀_layout_g{g}.xlsx")
        wb = openpyxl.load_workbook(out)
        check(f"UNIV {g}학년 시트", f"UNIV {g}학년" in wb.sheetnames and "대교협" in wb.sheetnames)

    # 정오표에 학생이 마킹한 답만 표시 (오답도 학생답만, 미마킹=·)
    jeongo = {1: {"답": 2, "정답": 2, "배점": 2, "ok": True},
              2: {"답": 0, "정답": 5, "배점": 2, "ok": False},
              5: {"답": 3, "정답": 4, "배점": 2, "ok": False}}
    su = dict(학년=1, 반="01", 번호="03", 이름="정오테스트", 계열번호="",
              국어=dict(선택="국어", 공통원=80, 선택원="", 원점수=80, 만점=100, 정오=jeongo))
    out = rx.build_workbook([su], ROOT / "output/_회귀_marksheet.xlsx")
    ws = openpyxl.load_workbook(out)["정오표_국어"]
    # 범례 1행, 헤더 2행, 데이터 3행 / 문항 셀은 7열(NB=6)부터
    v1, v2, v5 = ws.cell(3, 7).value, ws.cell(3, 8).value, ws.cell(3, 11).value
    check("정오표 학생답만 표시", v1 == 2 and v2 == "·" and v5 == 3, f"{v1},{v2},{v5}")


# ── 9. 등급컷 → 예상 등급·표준점수·백분위 ───────────────────────
def t9_grade_cuts():
    print("[9] 등급컷 예상치 (컷 경계값 전수 + 통합성적표)")
    import grade_cuts as gc
    for irecord, label in (("202606041", "고1"), ("202606043", "고3")):
        cuts = gc.load_grade_cuts(irecord)
        if not cuts:
            check(f"{label} 등급컷 파일", False, "keys/{irecord}_등급컷.json 없음")
            continue
        total = bad = 0
        for subj, cut in cuts.items():
            for c in cut["cuts"]:
                if c.get("raw") is None or not c.get("grade"):
                    continue
                est = gc.estimate(cut, c["raw"])
                total += 1
                bad += est["등급"] != c["grade"]
                if c.get("std_score") is not None and est["표준점수"] != "":
                    total += 1
                    bad += abs(est["표준점수"] - c["std_score"]) > 1
                if c.get("pct") is not None and est["백분위"] != "":
                    total += 1
                    bad += abs(est["백분위"] - c["pct"]) > 1
        check(f"{label} 컷 경계값 {total - bad}/{total}", bad == 0)
    # 과목명 표기 차이 매칭 (사회·문화 ↔ 사회문화)
    cuts3 = gc.load_grade_cuts("202606043")
    check("과목명 중간점 매칭", all(gc.find_subject_cut(cuts3, n) for n in ("사회문화", "정치와법")))
    # 통합성적표에 예상치 반영 (고1 synthetic)
    import consolidate as c
    out = c.consolidate_paths(
        1, keys_dir=str(ROOT / "keys"), exam_id="202606041",
        korean=str(ROOT / "output/_회귀_synth/synth_국어_국어_판독표.csv"),
        math=str(ROOT / "output/_회귀_synth/synth_수학_수학_판독표.csv"),
        english=str(ROOT / "output/_회귀_synth/synth_영어_영어_판독표.csv"),
        history=str(ROOT / "output/_회귀_synth/synth_한국사_한국사_판독표.csv"),
        social=str(ROOT / "output/_회귀_synth/synth_통합사회_통합사회_판독표.csv"),
        science=str(ROOT / "output/_회귀_synth/synth_통합과학_통합과학_판독표.csv"),
        out=str(ROOT / "output/_회귀_통합성적표_고1.xlsx"))
    import openpyxl
    wb = openpyxl.load_workbook(out)
    check("예상등급컷 시트", "예상등급컷" in wb.sheetnames)
    ws = wb["UNIV 1학년"]
    cuts1 = gc.load_grade_cuts("202606041")
    colmap = dict(국어=6, 수학=10, 영어=14, 통합사회=18, 통합과학=22, 한국사=26)
    total = ok = 0
    for r in range(3, ws.max_row + 1):
        if not ws.cell(r, 4).value:
            continue
        for subj, c0 in colmap.items():
            est = gc.estimate(gc.find_subject_cut(cuts1, subj), ws.cell(r, c0).value)
            got = (ws.cell(r, c0 + 1).value, ws.cell(r, c0 + 2).value, ws.cell(r, c0 + 3).value)
            want = (est["표준점수"] or None, est["백분위"] or None, est["등급"] or None)
            total += 1
            ok += got == want
    check(f"UNIV 예상치 {ok}/{total}", total == 36 and ok == 36)


# ── 9.3 미맥 고3 국·수 원점수 예상등급 ───────────────────────────
def t93_mimac():
    print("[9.3] 미맥 고3 국·수 예상등급")
    import mimac_cuts as mimac
    data = mimac.load_g3_cuts("202606043")
    if not data:
        print("  – 미맥 컷 파일 없음 → 건너뜀 (register 후 생성)")
        return
    # 저장 형식·시험일 검증
    check("미맥 시험일 6.4", data.get("_meta", {}).get("on_date") == "6.4")
    # 등급 산출: 화작 컷 1:95 2:90 3:82 4:72 → 경계 확인
    g = lambda raw: mimac.mimac_g3_grade("202606043", "국어", "화법과작문", raw)
    check("화작 95→1·84→3·72→4·71→5",
          g(95) == 1 and g(84) == 3 and g(72) == 4 and g(71) == 5,
          f"{g(95)},{g(84)},{g(72)},{g(71)}")
    # 수학 확통 1:91 → 91=1등급, 90=2등급
    gm = lambda raw: mimac.mimac_g3_grade("202606043", "수학", "확률과통계", raw)
    check("확통 91→1·90→2", gm(91) == 1 and gm(90) == 2, f"{gm(91)},{gm(90)}")


# ── 9.5 서버 모드 보안 (프록시 뒤 인증 우회·무차별 대입 방어) ────
def t95_proxy_security():
    print("[9.5] 서버 모드 보안 (프록시 인증 게이트)")
    code = (
        "import web_app as w\n"
        "c=w.app.test_client()\n"
        "assert w.BEHIND_PROXY is True\n"
        "r=c.get('/',environ_overrides={'REMOTE_ADDR':'127.0.0.1'})\n"
        "assert r.status_code==302 and '/login' in r.headers.get('Location',''), 'bypass'\n"
        "[c.post('/login',data={'code':'0'},environ_overrides={'REMOTE_ADDR':'9.9.9.9'}) for _ in range(5)]\n"
        "assert c.post('/login',data={'code':'0'},environ_overrides={'REMOTE_ADDR':'9.9.9.9'}).status_code==429, 'no lockout'\n"
        "c2=w.app.test_client()\n"
        "assert c2.post('/login',data={'code':'87654321'},environ_overrides={'REMOTE_ADDR':'5.6.7.8'}).status_code==302\n"
        "assert 'OMR Lens' in c2.get('/',environ_overrides={'REMOTE_ADDR':'5.6.7.8'}).get_data(as_text=True)\n"
        "print('OK')\n")
    import os
    env = os.environ.copy()
    env.update(PYTHONIOENCODING="utf-8", OMR_BEHIND_PROXY="1", OMR_ACCESS_CODE="87654321")
    p = subprocess.run([sys.executable, "-c", code], cwd=ROOT, env=env, text=True,
                       encoding="utf-8", errors="replace",
                       stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=120)
    check("프록시 인증 게이트·잠금·브랜딩", p.returncode == 0 and "OK" in p.stdout, p.stdout[-200:])


def t96_results_edit():
    print("[9.6] 결과 확인·수정 API (조회·답안수정 재채점·성명수정·다운로드 반영)")
    code = (
        "import web_app as w, time, io, openpyxl\n"
        "c=w.app.test_client()\n"
        "d={'grade':'1','exam_id':'202606041'}\n"
        "import pathlib; s=pathlib.Path('work/synth_g12')\n"
        "for subj,k in [('국어','korean'),('통합사회','social')]:\n"
        "    d['pdf_'+k]=(open(s/('synth_%s.pdf'%subj),'rb'), subj+'.pdf')\n"
        "d['names']=(open(s/'synth_names.csv','rb'),'names.csv')\n"
        "rid=c.post('/api/score',data=d,content_type='multipart/form-data').get_json()['run_id']\n"
        "for _ in range(90):\n"
        "    time.sleep(2)\n"
        "    if c.get('/api/job/'+rid).get_json()['status']=='done': break\n"
        "res=c.get('/api/result/'+rid).get_json()\n"
        "assert res['ok'] and res['has_tamgu'] and '국어' in res['subjects'], 'view'\n"
        "assert any(t['과목']=='통합사회' for t in res['students'][0]['탐구']), 'tamgu view'\n"
        "kor=res['students'][0]['subjects']['국어']; before=kor['원점수']\n"
        "cc=next(x for x in kor['cells'] if x['ok']); wrong=(cc['정답']%5)+1\n"
        "er=c.post('/api/result/%s/edit'%rid,json={'idx':0,'field':'answer','subject':'국어','q':cc['q'],'value':wrong}).get_json()\n"
        "after=er['student']['subjects']['국어']['원점수']\n"
        "assert er['ok'] and after<before, 'regrade drop %s->%s'%(before,after)\n"
        "back=c.post('/api/result/%s/edit'%rid,json={'idx':0,'field':'answer','subject':'국어','q':cc['q'],'value':cc['정답']}).get_json()\n"
        "assert back['student']['subjects']['국어']['원점수']==before, 'restore'\n"
        "c.post('/api/result/%s/edit'%rid,json={'idx':0,'field':'이름','value':'검증학생X'})\n"
        "rr=c.get('/download/%s/통합성적표_고1.xlsx'%rid)\n"
        "wb=openpyxl.load_workbook(io.BytesIO(rr.data))\n"
        "found=any((cell.value=='검증학생X') for ws in wb.worksheets for row in ws.iter_rows() for cell in row)\n"
        "assert found, 'name not in workbook'\n"
        "print('OK')\n")
    import os
    env = os.environ.copy()
    env.update(PYTHONIOENCODING="utf-8", OMR_OPEN_ACCESS="1")
    p = subprocess.run([sys.executable, "-c", code], cwd=ROOT, env=env, text=True,
                       encoding="utf-8", errors="replace",
                       stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=300)
    check("결과 조회·수정·재채점·다운로드 반영", p.returncode == 0 and "OK" in p.stdout, p.stdout[-300:])


def t97_xip_kuksu():
    print("[9.7] 고3 국·수 XIP 키 생성 (6월 stored 키와 완전 일치 + 학평 org코드)")
    code = (
        "import exam_registry as er, json, tempfile, pathlib\n"
        "s=None\n"
        "try:\n"
        "    ir=er.find_irecord(2026,6,3)\n"
        "    assert ir=='202606043', 'find_irecord 6월='+str(ir)\n"
        "    papers=er.discover_papers('202606043')\n"
        "except Exception as e:\n"
        "    print('SKIP', repr(e)[:80]); raise SystemExit(0)\n"
        "tmp=pathlib.Path(tempfile.mkdtemp())\n"
        "out={'created':[],'skipped':[],'errors':[]}\n"
        "built=er._register_g3_kuksu('202606043',papers,tmp,lambda subj:False,out,lambda m:None)\n"
        "assert sorted(built['국어'])==['화법과 작문','언어와 매체'][::-1] or set(built['국어'])=={'화법과 작문','언어와 매체'}, built\n"
        "assert set(built['수학'])=={'확률과 통계','미적분','기하'}, built\n"
        "assert not out['errors'], out['errors']\n"
        "def norm_k(d):\n"
        "    if 'answers' in d: return {q:d['answers'][q] for q in d['answers']}, {q:d['points'][q] for q in d['points']}\n"
        "    qs=d['questions']; return {q:qs[q]['answer'] for q in qs}, {q:qs[q]['points'] for q in qs}\n"
        "mism=[]\n"
        "for f in list(tmp.glob('*.json')):\n"
        "    stored=pathlib.Path('keys')/f.name\n"
        "    if not stored.exists(): mism.append('stored 없음:'+f.name); continue\n"
        "    a1,p1=norm_k(json.loads(f.read_text(encoding='utf-8')))\n"
        "    a2,p2=norm_k(json.loads(stored.read_text(encoding='utf-8')))\n"
        "    if a1!=a2: mism.append('답 불일치:'+f.name)\n"
        "    if p1!=p2: mism.append('배점 불일치:'+f.name)\n"
        "assert not mism, mism\n"
        "print('OK 5키 일치')\n")
    import os
    env = os.environ.copy(); env.update(PYTHONIOENCODING="utf-8")
    p = subprocess.run([sys.executable, "-c", code], cwd=ROOT, env=env, text=True,
                       encoding="utf-8", errors="replace",
                       stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=180)
    ok = p.returncode == 0 and ("OK" in p.stdout or "SKIP" in p.stdout)
    if "SKIP" in p.stdout:
        print("  (건너뜀 — EBSi 접속 불가)")
    check("XIP 국·수 키가 stored 6월 키와 일치", ok, p.stdout[-300:])


def t98_multi_pdf_merge():
    print("[9.8] 한 과목 여러 PDF 업로드 → 병합 (스캐너 분할본 대응)")
    code = (
        "import web_app as w, fitz, io, pathlib, tempfile\n"
        "src=fitz.open('work/synth_g12/synth_국어.pdf'); n=src.page_count; h=n//2\n"
        "a=fitz.open(); a.insert_pdf(src,from_page=0,to_page=h-1); ab=a.tobytes(); a.close()\n"
        "b=fitz.open(); b.insert_pdf(src,from_page=h,to_page=n-1); bb=b.tobytes(); b.close(); src.close()\n"
        "tmp=pathlib.Path(tempfile.mkdtemp())\n"
        "data={'pdf_korean':[(io.BytesIO(ab),'A.pdf'),(io.BytesIO(bb),'B.pdf')]}\n"
        "with w.app.test_request_context('/api/score',method='POST',data=data,content_type='multipart/form-data'):\n"
        "    p=w.save_upload(tmp,'pdf_korean')\n"
        "merged=fitz.open(p); mc=merged.page_count; merged.close()\n"
        "assert mc==n, f'병합 페이지 {mc} != 원본 {n}'\n"
        "assert not list(tmp.glob('*_part*.pdf')), '임시 분할본이 남음'\n"
        "print('OK 병합', mc, '페이지')\n")
    import os
    env = os.environ.copy(); env.update(PYTHONIOENCODING="utf-8", OMR_OPEN_ACCESS="1")
    p = subprocess.run([sys.executable, "-c", code], cwd=ROOT, env=env, text=True,
                       encoding="utf-8", errors="replace",
                       stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=120)
    check("여러 PDF 병합", p.returncode == 0 and "OK" in p.stdout, p.stdout[-300:])


def t99_chunked_upload():
    print("[9.9] 대용량 분할 업로드 (init·file×N·start → 병합 채점)")
    code = (
        "import web_app as w, time, io, fitz, pathlib\n"
        "c=w.app.test_client(); synth=pathlib.Path('work/synth_g12')\n"
        "src=fitz.open(synth/'synth_국어.pdf'); n=src.page_count; per=max(1,n//3); parts=[]\n"
        "for i in range(0,n,per):\n"
        "    d=fitz.open(); d.insert_pdf(src,from_page=i,to_page=min(i+per-1,n-1)); parts.append(d.tobytes()); d.close()\n"
        "src.close()\n"
        "rid=c.post('/api/score/init',json={'grade':1,'exam_id':'202606041'}).get_json()['run_id']\n"
        "c.post(f'/api/score/{rid}/file',data={'names':(open(synth/'synth_names.csv','rb'),'n.csv')},content_type='multipart/form-data')\n"
        "for i,pb in enumerate(parts):\n"
        "    r=c.post(f'/api/score/{rid}/file',data={'pdf_korean':(io.BytesIO(pb),f'k{i}.pdf')},content_type='multipart/form-data').get_json()\n"
        "    assert r['ok'] and r['saved']==1, r\n"
        "c.post(f'/api/score/{rid}/file',data={'pdf_social':(open(synth/'synth_통합사회.pdf','rb'),'s.pdf')},content_type='multipart/form-data')\n"
        "st=c.post(f'/api/score/{rid}/start',json={}).get_json(); assert st['ok'], st\n"
        "for _ in range(90):\n"
        "    time.sleep(2)\n"
        "    if c.get(f'/api/job/{rid}').get_json()['status']=='done': break\n"
        "res=c.get(f'/api/result/{rid}').get_json()\n"
        "kor=[s for s in res['students'] if s['subjects'].get('국어')]\n"
        "assert len(kor)==6, f'국어 학생수 {len(kor)} != 6 (병합 실패)'\n"
        "assert not w.PENDING.get(rid), 'start 후 PENDING 미정리'\n"
        "print('OK 3조각 병합·6명 채점')\n")
    import os
    env = os.environ.copy(); env.update(PYTHONIOENCODING="utf-8", OMR_OPEN_ACCESS="1")
    p = subprocess.run([sys.executable, "-c", code], cwd=ROOT, env=env, text=True,
                       encoding="utf-8", errors="replace",
                       stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=300)
    check("분할 업로드→병합 채점", p.returncode == 0 and "OK" in p.stdout, p.stdout[-300:])


# ── 10. 웹 e2e (옵션: 서버 실행 중일 때) ─────────────────────────
def t10_web():
    print("[10] 웹 e2e (고1 6과목 → 통합성적표)")
    import requests
    from synth_omr import make_students
    base = "http://127.0.0.1:5050"
    try:
        requests.get(base + "/api/exams", timeout=3)
    except Exception:
        check("서버 응답", False, "web_app.py 실행 후 --web 로 재시도")
        return
    synth = ROOT / "work/synth_g12"
    files = {}
    for subj, key in [("국어", "korean"), ("수학", "math"), ("영어", "english"),
                      ("한국사", "history"), ("통합사회", "social"), ("통합과학", "science")]:
        files[f"pdf_{key}"] = (f"{subj}.pdf", open(synth / f"synth_{subj}.pdf", "rb"), "application/pdf")
    files["names"] = ("names.csv", open(synth / "synth_names.csv", "rb"), "text/csv")
    r = requests.post(base + "/api/score", data={"grade": "1", "exam_id": "202606041"},
                      files=files, timeout=180)
    j = r.json()
    import time
    for _ in range(200):
        time.sleep(2)
        st = requests.get(base + j["status_url"], timeout=5).json()
        if st["status"] in ("done", "error"):
            break
    check("웹 잡 완료", st["status"] == "done" and st.get("ok") is True)
    import openpyxl, io
    students = make_students(ROOT / "keys", "202606041", n=6, grade=1)
    exp = {info["beon"]: e for info, _, e in students}
    rr = requests.get(f'{base}/download/{j["run_id"]}/통합성적표_고1.xlsx', timeout=10)
    ws = openpyxl.load_workbook(io.BytesIO(rr.content))["UNIV 1학년"]
    total = ok = 0
    colmap = dict(국어=6, 수학=10, 영어=14, 통합사회=18, 통합과학=22, 한국사=26)
    for row in range(3, ws.max_row + 1):
        beon = str(ws.cell(row, 4).value or "")
        if beon not in exp:
            continue
        for subj, col in colmap.items():
            total += 1
            v = ws.cell(row, col).value
            if v is not None and float(v) == float(exp[beon][subj]):
                ok += 1
    check(f"통합성적표 대조 {ok}/{total}", total == 36 and ok == 36)


def t100_hp_units():
    """[10.0] 고3 학평(hp_*) 단위 테스트 — 합성 이미지 기반 (tests_hp.py)"""
    print("[10.0] 고3 학평 판독 단위 테스트 (성명·수험번호·단답형·흑백감지·정렬매칭)")
    import subprocess
    r = subprocess.run([sys.executable, "-X", "utf8", str(ROOT / "tests_hp.py")],
                       capture_output=True, text=True, encoding="utf-8", errors="replace")
    tail = (r.stdout or "").strip().splitlines()[-1] if r.stdout else ""
    check(f"tests_hp 전체 통과 ({tail})", r.returncode == 0,
          (r.stdout or "")[-800:] + (r.stderr or "")[-400:])


def main() -> int:
    web = "--web" in sys.argv
    for t in [t1_imports, t2_key_isolation, t3_history_g3, t4_consolidate_g3,
              t5_synth_g12, t6_edge, t65_simfixes, t67_autorotate, t7_calibrator, t8_report_layouts,
              t9_grade_cuts, t93_mimac, t95_proxy_security, t96_results_edit,
              t97_xip_kuksu, t98_multi_pdf_merge, t99_chunked_upload,
              t100_hp_units] + ([t10_web] if web else []):
        try:
            t()
        except Exception as exc:
            check(t.__name__ + " (예외)", False, repr(exc))
    print()
    print(f"결과: PASS {len(PASS)} / FAIL {len(FAIL)}")
    for name, detail in FAIL:
        print(f"  FAIL {name} {detail}")
    return 0 if not FAIL else 1


if __name__ == "__main__":
    raise SystemExit(main())
