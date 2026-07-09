# -*- coding: utf-8 -*-
"""
OMR 채점 파이프라인 통합 실행기
score_lens / omr_scorer

  PDF(여러 학생) → 페이지별 판독 → 정답키 채점 → 엑셀/CSV 점수표

사용법
    python run.py input/scan.pdf --irecord 202606043
    python run.py input/scan.pdf --keys-dir keys --dpi 300 --debug

선행: keys/ 에 정답키 JSON 이 있어야 함 (ebsi_keys.py 로 생성).
"""
from __future__ import annotations
import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

import cv2

import omr_core as oc
import name_reader as nr
import id_reader as idr
from grade import load_keys, grade, normalize_elective


def write_excel(rows: list[dict], out_xlsx: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    # ── 점수표 ───────────────────────────────────────────
    ws = wb.active
    ws.title = "점수표"
    head = ["페이지", "학교(학원)번호", "반", "번호", "성명", "선택과목",
            "원점수", "만점", "맞은개수", "틀린문항", "미마킹", "중복마킹"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col, w in zip("ABCDEFGHIJKL",
                      [6, 13, 5, 6, 10, 9, 7, 6, 9, 28, 16, 12]):
        ws.column_dimensions[col].width = w
    center = Alignment(horizontal="center", vertical="center")
    for r, row in enumerate(rows, start=2):
        ws.cell(r, 1, row["page"]).alignment = center
        ws.cell(r, 2, row.get("school", "")).alignment = center
        ws.cell(r, 3, row.get("ban", "")).alignment = center
        ws.cell(r, 4, row.get("beon", "")).alignment = center
        ws.cell(r, 5, row.get("name", "")).alignment = center
        ws.cell(r, 6, row["elective"] or "?").alignment = center
        ws.cell(r, 7, row["score"] if row["score"] is not None else "")
        ws.cell(r, 8, row["max_score"])
        ws.cell(r, 9, row["correct_count"])
        ws.cell(r, 10, ", ".join(map(str, row["wrong"])))
        ws.cell(r, 11, ", ".join(map(str, row["blank"])))
        ws.cell(r, 12, ", ".join(map(str, row["dup"])))
    ws.freeze_panes = "A2"

    # ── 문항별 정오 ──────────────────────────────────────
    ws2 = wb.create_sheet("문항별정오")
    base_cols = ["페이지", "반", "번호", "성명", "선택과목", "점수"]
    nb = len(base_cols)
    ws2.append(base_cols + [str(q) for q in range(1, 46)])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    gray = PatternFill("solid", fgColor="D9D9D9")
    for r, row in enumerate(rows, start=2):
        ws2.cell(r, 1, row["page"])
        ws2.cell(r, 2, row.get("ban", ""))
        ws2.cell(r, 3, row.get("beon", ""))
        ws2.cell(r, 4, row.get("name", ""))
        ws2.cell(r, 5, row["elective"] or "?")
        ws2.cell(r, 6, row["score"] if row["score"] is not None else "")
        for q in range(1, 46):
            cell = ws2.cell(r, nb + q)
            pq = row["per_q"].get(q)
            if not pq:
                continue
            stu = pq["student"]
            cell.value = {0: "·", -1: "중"}.get(stu, stu)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = green if pq["result"] == "O" else red if pq["result"] == "X" else gray
    ws2.freeze_panes = ws2.cell(1, nb + 1).coordinate
    ws2.column_dimensions["D"].width = 9
    for q in range(1, 46):
        ws2.column_dimensions[ws2.cell(1, nb + q).column_letter].width = 3.5

    wb.save(out_xlsx)


def main() -> int:
    ap = argparse.ArgumentParser(description="OMR 채점 → 점수표")
    ap.add_argument("pdf", type=Path)
    ap.add_argument("--irecord", default=None, help="정답키 선택용 EBSi irecord")
    ap.add_argument("--keys-dir", type=Path, default=Path("keys"))
    ap.add_argument("--dpi", type=int, default=300)
    ap.add_argument("--out", type=Path, default=Path("output"))
    ap.add_argument("--debug", action="store_true", help="페이지별 판독 오버레이 저장")
    args = ap.parse_args()

    if not args.pdf.exists():
        print(f"[오류] PDF 없음: {args.pdf}", file=sys.stderr); return 1
    keys = load_keys(args.keys_dir, args.irecord)
    if not keys:
        print(f"[오류] 정답키 없음: {args.keys_dir} (ebsi_keys.py 로 먼저 생성)", file=sys.stderr)
        return 1
    print(f"정답키 로드: {list(keys)}")

    args.out.mkdir(parents=True, exist_ok=True)

    # PNG 파일을 만들지 않고 PDF → 메모리(ndarray) 직행으로 렌더한다.
    n_pages = oc.pdf_page_count(args.pdf)
    print(f"{n_pages}페이지 ({args.dpi} DPI, 메모리 직행)")

    print("격자 보정(여러 페이지 중앙값 합의)...")
    n_sample = min(n_pages, 15)
    step = n_pages / n_sample
    sample_idx = [int(k * step) for k in range(n_sample)]
    template = oc.calibrate_template(
        (im for _, im in oc.iter_pdf_pages(args.pdf, args.dpi, pages=sample_idx)))
    print(f"  y0={template['y0']:.0f} pitch={template['pitch']:.1f} "
          f"(보정 {template['n_pages']}페이지)")

    rows, flagged = [], []
    for pi, raw in oc.iter_pdf_pages(args.pdf, args.dpi):
        i = pi + 1
        img = oc.load_page(raw)          # auto_orient 적용 (경로 로드와 동일 불변식)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        res = oc.process_page(img, template)
        name = nr.read_name(gray)
        sid = idr.read_id(gray)
        elective = normalize_elective(res["subject"])
        key = keys.get(elective)
        base = dict(page=i, name=name, school=sid["school"], ban=sid["ban"],
                    beon=sid["beon"], elective=res["subject"])
        if key:
            g = grade(res["answers"], key)
            row = dict(**base, score=g["score"], max_score=g["max_score"],
                       correct_count=g["correct_count"], wrong=g["wrong"],
                       blank=g["blank"], dup=g["dup"], per_q=g["per_q"])
        else:
            row = dict(**base, score=None, max_score=100, correct_count=0,
                       wrong=[], blank=[], dup=[], per_q={})
        rows.append(row)
        # 신뢰도 점검: 학교번호 이상 / 선택과목 미검출 / 미마킹·중복 과다 → 수동 확인
        warns = []
        if sid["school"] != "19718":
            warns.append(f"학교번호 {sid['school']}")
        if not key:
            warns.append("선택과목 미검출/키없음")
        if len(row["blank"]) > 8:
            warns.append(f"미마킹 {len(row['blank'])}개")
        if row["dup"]:
            warns.append(f"중복마킹 {len(row['dup'])}개")
        if warns:
            flagged.append((i, warns))
        flag = ("  ⚠️ " + ", ".join(warns)) if warns else ""
        sc = row["score"] if row["score"] is not None else "-"
        print(f"  p{i:>3}: {name:<5} {sid['ban']}-{sid['beon']} "
              f"{res['subject'] or '?':<8} {sc}/{row['max_score']}점{flag}")

    def _save(path: Path, saver) -> Path:
        # 결과 파일이 엑셀 등에서 열려 있으면 PermissionError → 시각을 붙인 대체 이름으로 저장
        try:
            saver(path)
            return path
        except PermissionError:
            alt = path.with_name(f"{path.stem}_{datetime.now():%H%M%S}{path.suffix}")
            print(f"  ⚠️ 파일이 열려 있어 대체 이름으로 저장: {alt.name}")
            saver(alt)
            return alt

    # CSV (문항별 학생답 1~45 포함 — 통합 정오표에서 학생 답안 표시용)
    def _write_csv(path: Path) -> None:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["페이지", "학교(학원)번호", "반", "번호", "성명", "선택과목",
                        "원점수", "만점", "맞은개수", "틀린문항", "미마킹", "중복마킹"]
                       + [str(q) for q in range(1, 46)])
            for r in rows:
                pq = r.get("per_q", {})
                w.writerow([r["page"], r.get("school", ""), r.get("ban", ""), r.get("beon", ""),
                            r.get("name", ""), r["elective"] or "?", r["score"], r["max_score"],
                            r["correct_count"], " ".join(map(str, r["wrong"])),
                            " ".join(map(str, r["blank"])), " ".join(map(str, r["dup"]))]
                           + [pq.get(q, {}).get("student", "") for q in range(1, 46)])

    csv_path = _save(args.out / f"{args.pdf.stem}_점수표.csv", _write_csv)
    xlsx_path = _save(args.out / f"{args.pdf.stem}_점수표.xlsx",
                      lambda p: write_excel(rows, p))

    scored = [r["score"] for r in rows if r["score"] is not None]
    if scored:
        print(f"\n응시 {len(scored)}명  평균 {sum(scored)/len(scored):.1f}점  "
              f"최고 {max(scored)}  최저 {min(scored)}")
    if flagged:
        print(f"⚠️ 수동 확인 권고 {len(flagged)}건: "
              + ", ".join(f"p{p}({'/'.join(w)})" for p, w in flagged))
    print(f"\n완료:\n  {xlsx_path}\n  {csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
