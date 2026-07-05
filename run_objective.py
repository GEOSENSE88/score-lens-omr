# -*- coding: utf-8 -*-
"""
범용 객관식 OMR 채점 파이프라인 (score_lens / omr_scorer)

단일 과목·N문항 객관식 답안지(예: 고1·2 영어45 / 한국사20 / 통합사회25 / 통합과학25)를
하나의 스크립트로 처리한다. 양식 좌표는 --template(JSON), 정답·배점은 keys/ 에서
--subject 로 매칭한다. 수험번호는 --id-layout 으로 카드 양식(g3 / g12)을 고른다.

    python run_objective.py "D:/scan/통합과학.pdf" --subject 통합과학 \
        --template templates/science_g12.json --irecord <고1·2 exam_id> \
        --id-layout g12 --names output/국어_판독표.csv --out output

정답키가 아직 없으면 판독표(마킹 값)만 생성한다.
"""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

import cv2
import fitz

import omr_core as oc
import history_core as objreader   # read_answers 가 template 기반·과목 무관
import id_reader as idr
from run_math import load_name_map


def render_pages(pdf: Path, dpi: int, out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(pdf)
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    paths = []
    for i, page in enumerate(doc, 1):
        p = out_dir / f"{pdf.stem}_p{i:03d}.png"
        page.get_pixmap(matrix=mat, alpha=False).save(p)
        paths.append(p)
    doc.close()
    return paths


def load_template(path: Path) -> dict:
    tmpl = json.loads(path.read_text(encoding="utf-8"))
    tmpl["objective"] = [dict(b) for b in tmpl["objective"]]
    if tmpl.get("short"):                                  # 수학 단답형 (백·십·일)
        tmpl["short"] = {int(q): dict(v) for q, v in tmpl["short"].items()}
    return tmpl


def n_questions(template: dict) -> int:
    nq = max(b["q0"] + b["nrows"] - 1 for b in template["objective"])
    if template.get("short"):
        nq = max(nq, max(template["short"]))
    return nq


def read_all(gray, xf, template: dict) -> dict[int, int]:
    """객관식 + (있으면) 수학 단답형 판독."""
    answers = objreader.read_answers(gray, xf, template)
    if template.get("short"):
        import math_core as mc
        answers.update(mc.read_short(gray, xf, mc._normalize_template(template)))
    return answers


def load_subject_key(keys_dir: Path, subject: str, exam_id: str | None) -> dict | None:
    cands = []
    for p in keys_dir.glob("*.json"):
        try:
            k = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        if k.get("subject") != subject:
            continue
        if exam_id and str(k.get("exam_id") or k.get("irecord")) != str(exam_id):
            continue
        cands.append(k)
    if not cands:
        return None
    cands.sort(key=lambda k: (not bool(k.get("complete")), -int(k.get("item_count", 0))))
    return cands[0]


def grade_known(answers: dict[int, int], key: dict | None, nq: int) -> dict:
    if not key:
        return dict(score=None, max_score=0, correct=[], wrong=[], blank=[], dup=[], per_q={})
    kans = {int(q): int(v) for q, v in key["answers"].items()}
    kpts = {int(q): float(v) for q, v in key["points"].items()}   # 고1·2 통합과목 1.5점 등 소수 배점
    score, correct, wrong, blank, dup, per_q = 0.0, [], [], [], [], {}
    for q in sorted(kans):
        stu = answers.get(q, 0)
        ok = stu == kans[q]
        if stu == 0:
            blank.append(q)
        elif stu == -1:
            dup.append(q)
        elif ok:
            score += kpts[q]
            correct.append(q)
        else:
            wrong.append(q)
        per_q[q] = {"student": stu, "answer": kans[q], "points": kpts[q], "ok": ok}

    def _clean(v: float):
        return int(v) if float(v).is_integer() else round(v, 1)

    return dict(score=_clean(score), max_score=_clean(sum(kpts.values())), correct=correct,
                wrong=wrong, blank=blank, dup=dup, per_q=per_q)


def write_outputs(rows, out: Path, stem: str, subject: str, key, nq: int):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    out.mkdir(parents=True, exist_ok=True)
    headers = (["페이지", "학년/학교", "반", "번호", "성명", "점수", "만점",
                "틀린문항", "미마킹", "중복마킹", "확인필요"]
               + [str(q) for q in range(1, nq + 1)])

    csv_path = out / f"{stem}_{subject}_판독표.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow(
                [r["page"], r["school"], r["ban"], r["beon"], r["name"], r["score"], r["max_score"],
                 " ".join(map(str, r["wrong"])), " ".join(map(str, r["blank"])),
                 " ".join(map(str, r["dup"])), r.get("warn", "")]
                + [r["answers"].get(q, "") for q in range(1, nq + 1)]
            )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{subject} 판독"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    gray = PatternFill("solid", fgColor="D9D9D9")
    amber = PatternFill("solid", fgColor="FFE699")
    known = set(int(q) for q in (key or {}).get("answers", {}))
    center = Alignment(horizontal="center")
    nb = 11   # 문항 앞 기본 컬럼 수
    for row_idx, r in enumerate(rows, 2):
        base = [r["page"], r["school"], r["ban"], r["beon"], r["name"], r["score"], r["max_score"],
                " ".join(map(str, r["wrong"])), " ".join(map(str, r["blank"])),
                " ".join(map(str, r["dup"])), r.get("warn", "")]
        for col_idx, value in enumerate(base, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = center
            if col_idx == nb and value:          # 확인필요 강조
                cell.fill = amber
        for q in range(1, nq + 1):
            cell = ws.cell(row_idx, nb + q, r["answers"].get(q, ""))
            cell.alignment = center
            if q in known:
                cell.fill = green if q not in r["wrong"] and q not in r["blank"] and q not in r["dup"] else red
            else:
                cell.fill = gray
    ws.freeze_panes = ws.cell(2, nb + 1).coordinate
    for col in range(1, nb + nq + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 4.5 if col > nb else 10
    xlsx_path = out / f"{stem}_{subject}_판독표.xlsx"
    wb.save(xlsx_path)
    return xlsx_path, csv_path


def main() -> int:
    ap = argparse.ArgumentParser(description="범용 객관식 OMR 채점 → 판독표")
    ap.add_argument("pdf", type=Path)
    ap.add_argument("--subject", required=True, help="과목명(키 매칭·출력 라벨): 영어/한국사/통합사회/통합과학 등")
    ap.add_argument("--template", type=Path, required=True, help="객관식 좌표 템플릿 JSON")
    ap.add_argument("--irecord", default=None)
    ap.add_argument("--keys-dir", type=Path, default=Path("keys"))
    ap.add_argument("--names", default=None, help="성명 조인용 CSV")
    ap.add_argument("--id-layout", default="g3", choices=list(idr.LAYOUTS), help="수험번호 카드 양식")
    ap.add_argument("--dpi", type=int, default=300)
    ap.add_argument("--out", type=Path, default=Path("output"))
    args = ap.parse_args()

    template = load_template(args.template)
    nq = n_questions(template)
    key = load_subject_key(args.keys_dir, args.subject, args.irecord)
    if key:
        print(f"{args.subject} 키: {key.get('item_count', len(key.get('answers', {})))}문항, complete={key.get('complete', False)}")
    else:
        print(f"{args.subject} 키 없음: 판독표만 생성")
    names_full, names_beon = load_name_map(args.names)

    # ⚠️ 렌더 폴더는 --out 아래에 둔다 — 공용 work/ 를 쓰면 웹에서 같은 과목
    #    잡 2개가 동시에 돌 때 페이지 PNG 가 서로 덮여 학생이 뒤섞인다.
    work = args.out / f"_pages_{args.pdf.stem}_{args.subject}"
    pages = render_pages(args.pdf, args.dpi, work)
    if not pages:
        print(f"[오류] PDF 에 페이지가 없습니다: {args.pdf}", file=sys.stderr)
        return 1
    rows = []
    for i, pg in enumerate(pages, 1):
        img = oc.load_page(str(pg))
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        anchor_ok = oc.detect_top_anchor(gray) is not None
        xf = oc.make_anchor_transform(gray)
        sid = idr.read_id(gray, args.id_layout, xf)
        answers = read_all(gray, xf, template)
        g = grade_known(answers, key, nq)
        ident = sid.get("grade", "") or sid.get("school", "")
        name = names_full.get((sid["ban"], sid["beon"])) or names_beon.get(sid["beon"], "")
        # 수동 확인 경고: 앵커 미검출 / 수험번호 오판독 / 미마킹 다수 / 중복마킹
        warns = []
        if not anchor_ok:
            warns.append("앵커미검출")   # 상단 타이밍마크 미검출 → 좌표 보정 없이 판독됨
        if "?" in f"{ident}{sid['ban']}{sid['beon']}":
            warns.append("수험번호")
        if len(g["blank"]) >= 5:
            warns.append(f"미마킹{len(g['blank'])}")
        if g["dup"]:
            warns.append(f"중복{len(g['dup'])}")
        warn = " ".join(warns)
        rows.append(dict(page=i, school=ident, ban=sid["ban"], beon=sid["beon"],
                         name=name, answers=answers, warn=warn, **g))
        sc = g["score"] if g["score"] is not None else "-"
        flag = f"  ⚠️ {warn}" if warn else ""
        print(f"p{i:02d} {name:<5} {sid['ban']}-{sid['beon']} score {sc}/{g['max_score']}{flag}")

    # 배치 수준 이상 감지: 키가 있는데 평균 점수가 비정상적으로 낮으면
    # 과목/템플릿/카드 불일치(예: 영어 카드를 국어로 채점) 가능성이 크다.
    # 평균 계산에서는 '사실상 백지(결시)' 행만 제외한다 — 일반 경고 행을 다
    # 빼면 바꿔치기 케이스(전원 미마킹 경고)가 감지망을 빠져나간다.
    scored = [r["score"] for r in rows
              if r["score"] is not None and len(r["blank"]) < nq * 0.8]
    if key and len(scored) >= 3:
        max_sc = max((r["max_score"] for r in rows), default=0) or 0
        if max_sc and (sum(scored) / len(scored)) < 0.25 * max_sc:
            print(f"⚠️⚠️ 평균 점수 {sum(scored)/len(scored):.1f}/{max_sc} — 비정상적으로 낮습니다. "
                  f"과목·템플릿·답안지 종류가 맞는지 확인하세요 (--subject {args.subject}).")
            for r in rows:
                r["warn"] = (r.get("warn", "") + " 평균저점").strip()

    xlsx, csvp = write_outputs(rows, args.out, args.pdf.stem, args.subject, key, nq)
    print(f"완료:\n  {xlsx}\n  {csvp}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
