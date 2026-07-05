# -*- coding: utf-8 -*-
"""Inquiry OMR pipeline: selected subject codes + answers -> subject-specific scoring."""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

import cv2
import fitz

import ebsi_xip_keys
import explore_core as xc
from run_math import load_name_map


EXPLORE_ID_COLS = [139, 201, 265, 322, 387, 506, 562, 627, 680]
EXPLORE_ID_Y0 = 1285.0
EXPLORE_ID_PITCH = 100.0


# ⚠️ CODE_INFO 의 paper_id 는 아래 회차(2026-06 학평) 전용 — 다른 회차는 자동 다운로드 불가
CODE_INFO_EXAM_ID = "202606043"

CODE_INFO = {
    11: ("생활과윤리", "27037325"),
    12: ("윤리와사상", "27037326"),
    13: ("한국지리", "27037327"),
    14: ("세계지리", "27037328"),
    15: ("동아시아사", "27037329"),
    16: ("세계사", "27037330"),
    17: ("경제", "27037332"),
    18: ("정치와법", "27037331"),
    19: ("사회문화", "27037333"),
    20: ("물리학Ⅰ", "27037334"),
    21: ("화학Ⅰ", "27037336"),
    22: ("생명과학Ⅰ", "27037338"),
    23: ("지구과학Ⅰ", "27037340"),
    24: ("물리학Ⅱ", "27037335"),
    25: ("화학Ⅱ", "27037337"),
    26: ("생명과학Ⅱ", "27037339"),
    27: ("지구과학Ⅱ", "27037341"),
}


def read_explore_id(gray) -> dict:
    def digit(cx):
        vals = [xc.oc._darkness(gray, cx, EXPLORE_ID_Y0 + EXPLORE_ID_PITCH * r, 13) for r in range(10)]
        i = max(range(10), key=lambda idx: vals[idx])
        return i if vals[i] >= 50 else -1

    digs = [digit(cx) for cx in EXPLORE_ID_COLS]

    def join(seq):
        return "".join(str(x) if x >= 0 else "?" for x in seq)

    return {"school": join(digs[:5]), "ban": join(digs[5:7]), "beon": join(digs[7:9])}


def render_pages(pdf: Path, dpi: int, out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(pdf)
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    paths = []
    for i, page in enumerate(doc, 1):
        p = out_dir / f"{pdf.stem}_p{i:03d}.png"
        page.get_pixmap(matrix=mat, alpha=False).save(p)
        paths.append(p)
    doc.close()
    return paths


def load_key(keys_dir: Path, exam_id: str | None, subject: str) -> dict | None:
    for p in keys_dir.glob("*.json"):
        k = json.loads(p.read_text(encoding="utf-8"))
        if exam_id and str(k.get("exam_id") or k.get("irecord")) != str(exam_id):
            continue
        if k.get("subject") == subject and k.get("complete"):
            return k
    return None


def ensure_key(keys_dir: Path, exam_id: str | None, code: int) -> dict | None:
    info = CODE_INFO.get(code)
    if not info:
        return None
    subject, paper_id = info
    key = load_key(keys_dir, exam_id, subject)
    if key or not exam_id:
        return key
    # 자동 다운로드는 CODE_INFO 의 paper_id 가 고정된 회차(6월)만 안전함
    if str(exam_id) != CODE_INFO_EXAM_ID:
        raise SystemExit(
            f"[정답키 없음] {subject} — 자동 다운로드는 6월 회차({CODE_INFO_EXAM_ID}) paper_id 에 "
            f"고정되어 있습니다. 요청 회차({exam_id})를 채점하려면 CODE_INFO 의 paper_id 를 "
            "새 회차로 갱신하거나 키 JSON 을 keys/ 에 수동으로 준비하세요."
        )
    ebsi_xip_keys.write_key(str(exam_id), subject, paper_id, keys_dir)
    return load_key(keys_dir, exam_id, subject)


def grade_known(answers: dict[int, int], key: dict | None) -> dict:
    if not key:
        return dict(score=None, max_score=0, correct=[], wrong=[], blank=[], dup=[], per_q={})
    kans = {int(q): int(v) for q, v in key["answers"].items()}
    kpts = {int(q): int(v) for q, v in key["points"].items()}
    score, correct, wrong, blank, dup, per_q = 0, [], [], [], [], {}
    for q in sorted(kans):
        stu = answers.get(q, 0)
        ok = stu == kans[q]
        if stu == 0:
            blank.append(q)
        elif stu == -1:
            dup.append(q)
        elif ok:
            score += kpts[q]
            correct.append(q)
        else:
            wrong.append(q)
        per_q[q] = {"student": stu, "answer": kans[q], "points": kpts[q], "ok": ok}
    return dict(score=score, max_score=sum(kpts.values()), correct=correct, wrong=wrong, blank=blank, dup=dup, per_q=per_q)


def write_outputs(rows: list[dict], out: Path, stem: str) -> tuple[Path, Path]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    out.mkdir(parents=True, exist_ok=True)
    base_headers = ["페이지", "학교번호", "반", "번호", "성명", "총점", "총점만점"]
    sel_headers = []
    for label in ["제1선택", "제2선택"]:
        sel_headers += [f"{label}코드", f"{label}과목", f"{label}점수", f"{label}만점", f"{label}틀린문항"]
        sel_headers += [f"{label}_{q}" for q in range(1, 21)]
    headers = base_headers + sel_headers

    csv_path = out / f"{stem}_탐구_판독표.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            values = [r["page"], r["school"], r["ban"], r["beon"], r["name"], r["total_score"], r["total_max"]]
            for sel in ["sel1", "sel2"]:
                s = r[sel]
                values += [s["code"], s["subject"], s["score"], s["max_score"], " ".join(map(str, s["wrong"]))]
                values += [s["answers"].get(q, "") for q in range(1, 21)]
            w.writerow(values)

    wb = Workbook()
    ws = wb.active
    ws.title = "탐구 판독"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    gray = PatternFill("solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center")
    for row_idx, r in enumerate(rows, 2):
        flat = [r["page"], r["school"], r["ban"], r["beon"], r["name"], r["total_score"], r["total_max"]]
        col = 1
        for value in flat:
            ws.cell(row_idx, col, value).alignment = center
            col += 1
        for sel in ["sel1", "sel2"]:
            s = r[sel]
            for value in [s["code"], s["subject"], s["score"], s["max_score"], " ".join(map(str, s["wrong"]))]:
                ws.cell(row_idx, col, value).alignment = center
                col += 1
            known = set(s["per_q"])
            for q in range(1, 21):
                cell = ws.cell(row_idx, col, s["answers"].get(q, ""))
                cell.alignment = center
                if q in known:
                    cell.fill = green if q not in s["wrong"] and q not in s["blank"] and q not in s["dup"] else red
                else:
                    cell.fill = gray
                col += 1
    ws.freeze_panes = "H2"
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 4.5 if col > 12 else 10
    xlsx_path = out / f"{stem}_탐구_판독표.xlsx"
    wb.save(xlsx_path)
    return xlsx_path, csv_path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("pdf", type=Path)
    ap.add_argument("--irecord", default=None)
    ap.add_argument("--keys-dir", type=Path, default=Path("keys"))
    ap.add_argument("--names", default=None, help="Existing score CSV for name join")
    ap.add_argument("--template", type=Path, default=None)
    ap.add_argument("--dpi", type=int, default=300)
    ap.add_argument("--out", type=Path, default=Path("output"))
    args = ap.parse_args()

    template = xc.load_template(args.template)
    names_full, names_beon = load_name_map(args.names)
    work = Path("work") / f"{args.pdf.stem}_explore"
    pages = render_pages(args.pdf, args.dpi, work)
    rows = []

    for i, pg in enumerate(pages, 1):
        img = xc.oc.load_page(str(pg))
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        sid = read_explore_id(gray)
        res = xc.process_page(str(pg), template)
        row = {
            "page": i,
            "school": sid["school"],
            "ban": sid["ban"],
            "beon": sid["beon"],
            "name": names_full.get((sid["ban"], sid["beon"])) or names_beon.get(sid["beon"], ""),
        }
        total_score = 0
        total_max = 0
        for sel in ["sel1", "sel2"]:
            code = res["selections"][sel]["code"]
            subject = CODE_INFO.get(code, ("", ""))[0] if code else ""
            key = ensure_key(args.keys_dir, args.irecord, code) if code else None
            g = grade_known(res["selections"][sel]["answers"], key)
            row[sel] = {"code": code, "subject": subject, "answers": res["selections"][sel]["answers"], **g}
            total_score += g["score"] or 0
            total_max += g["max_score"] or 0
        row["total_score"] = total_score
        row["total_max"] = total_max
        rows.append(row)
        print(
            f"p{i:02d} {row['name']:<5} {sid['ban']}-{sid['beon']} "
            f"{row['sel1']['subject']} {row['sel1']['score']}/{row['sel1']['max_score']} + "
            f"{row['sel2']['subject']} {row['sel2']['score']}/{row['sel2']['max_score']} = "
            f"{row['total_score']}/{row['total_max']}"
        )

    xlsx, csvp = write_outputs(rows, args.out, args.pdf.stem)
    print(f"완료:\n  {xlsx}\n  {csvp}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
