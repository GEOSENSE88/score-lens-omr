# -*- coding: utf-8 -*-
"""
수학 OMR 채점 파이프라인
score_lens / omr_scorer

  PDF(수학 31명) → 판독(객관식+단답형+선택과목) → 정답키 채점 → 엑셀/CSV

학생 식별: 수험번호(반/번호) 자동판독. 성명은 국어 점수표 CSV에서 (반,번호)로 조인.
사용법:
    python run_math.py "D:\\scan\\수학.pdf" --irecord 202606043 --names output/국어_점수표.csv
"""
from __future__ import annotations
import argparse, csv, sys
from pathlib import Path
import cv2

import math_core as mc
import id_reader as idr
from grade import load_keys, normalize_elective


def load_name_map(csv_path):
    """국어 점수표 CSV → ({(반,번호): 성명}, {번호: 성명}). 번호는 폴백 조인용.

    CSV 가 아니거나 읽기 실패 시 크래시 대신 경고 후 빈 매핑 반환.
    """
    full, bybeon = {}, {}
    if not (csv_path and Path(csv_path).exists()):
        return full, bybeon
    try:
        with open(csv_path, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                nm = (row.get("성명") or "").strip()
                if not nm:
                    continue
                ban, beon = row.get("반", "").strip(), row.get("번호", "").strip()
                full[(ban, beon)] = nm
                bybeon[beon] = "" if beon in bybeon and bybeon[beon] != nm else nm  # 충돌시 비움
    except (UnicodeDecodeError, csv.Error, OSError) as exc:
        print(f"⚠️ 성명 CSV를 읽을 수 없어 무시합니다 ({Path(csv_path).name}): {exc}")
        return {}, {}
    return full, bybeon


def grade_math(answers, key):
    q = key["questions"]
    score, wrong, blank = 0, [], []
    for n in range(1, 31):
        kk = q[str(n)]; stu = answers.get(n, 0)
        if kk["type"] == "short":
            ok = (stu == kk["answer"])
            if stu == 0:
                blank.append(n)
        else:
            ok = (stu == kk["answer"])
            if stu == 0:
                blank.append(n)
        if ok:
            score += kk["points"]
        else:
            wrong.append(n)
    return dict(score=score, wrong=wrong, blank=blank)


def write_outputs(rows, out, stem):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "점수표"
    head = ["페이지", "학교(학원)번호", "반", "번호", "성명", "선택과목",
            "원점수", "만점", "맞은문항수", "틀린문항"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="375623")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col, w in zip("ABCDEFGHIJ", [6, 13, 5, 6, 10, 11, 7, 6, 9, 30]):
        ws.column_dimensions[col].width = w
    ctr = Alignment(horizontal="center")
    for r, row in enumerate(rows, 2):
        for c, v in enumerate([row["page"], row["school"], row["ban"], row["beon"],
                               row["name"], row["elective"] or "?", row["score"],
                               row["max"], 30 - len(row["wrong"]),
                               ", ".join(map(str, row["wrong"]))], 1):
            cell = ws.cell(r, c, v)
            if c <= 8:
                cell.alignment = ctr
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("문항별정오")
    ws2.append(["페이지", "반", "번호", "성명", "선택과목", "점수"] + [str(q) for q in range(1, 31)])
    for c in ws2[1]:
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    green = PatternFill("solid", fgColor="C6EFCE"); red = PatternFill("solid", fgColor="FFC7CE")
    for r, row in enumerate(rows, 2):
        for c, v in enumerate([row["page"], row["ban"], row["beon"], row["name"],
                               row["elective"] or "?", row["score"]], 1):
            ws2.cell(r, c, v)
        ws_wrong = set(row["wrong"])
        for n in range(1, 31):
            cell = ws2.cell(r, 6 + n, row["answers"].get(n, ""))
            cell.alignment = Alignment(horizontal="center")
            cell.fill = red if n in ws_wrong else green
    ws2.freeze_panes = "G2"
    for n in range(1, 31):
        ws2.column_dimensions[ws2.cell(1, 6 + n).column_letter].width = 4.5

    def _save(base, ext, saver):
        for cand in [out / f"{base}{ext}"] + [out / f"{base}_{k}{ext}" for k in range(2, 9)]:
            try:
                saver(cand); return cand
            except PermissionError:
                print(f"  (열려있어 건너뜀: {cand.name})")
        raise PermissionError(base)
    xlsx = _save(f"{stem}_수학_점수표", ".xlsx", wb.save)

    def _csv(path):
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["페이지", "학교번호", "반", "번호", "성명", "선택과목", "원점수", "만점", "틀린문항"]
                       + [str(q) for q in range(1, 31)] + ["확인필요"])
            for row in rows:
                ans = row.get("answers", {})
                w.writerow([row["page"], row["school"], row["ban"], row["beon"], row["name"],
                            row["elective"] or "?", row["score"], row["max"],
                            " ".join(map(str, row["wrong"]))]
                           + [ans.get(q, "") for q in range(1, 31)] + [row.get("warn", "")])
    csvp = _save(f"{stem}_수학_점수표", ".csv", _csv)
    return xlsx, csvp


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("pdf", type=Path)
    ap.add_argument("--irecord", default=None)
    ap.add_argument("--keys-dir", type=Path, default=Path("keys"))
    ap.add_argument("--names", default=None, help="국어 점수표 CSV (성명 조인용)")
    ap.add_argument("--template", type=Path, default=None, help="수학 OMR 좌표 템플릿 JSON")
    ap.add_argument("--dpi", type=int, default=300)
    ap.add_argument("--out", type=Path, default=Path("output"))
    args = ap.parse_args()
    if not args.pdf.exists():
        print(f"[오류] PDF 없음: {args.pdf}"); return 1

    import json
    keys = {}
    for p in args.keys_dir.glob("*수학*.json"):
        k = json.loads(p.read_text(encoding="utf-8"))
        if not args.irecord or str(k.get("irecord")) == str(args.irecord):
            keys[normalize_elective(k["elective"])] = k
    if not keys:
        print("[오류] 수학 정답키 없음 (math_keys.py 로 생성)"); return 1
    print(f"수학 정답키: {list(keys)}")
    names_full, names_beon = load_name_map(args.names)
    print(f"성명 조인: {len(names_full)}명 (국어 점수표)")
    template = mc.load_template(args.template)
    if args.template:
        print(f"템플릿: {args.template}")

    print(f"{mc.oc.pdf_page_count(args.pdf)}페이지 처리...")

    rows = []
    for pi, raw in mc.oc.iter_pdf_pages(args.pdf, args.dpi):
        i = pi + 1
        img = mc.oc.load_page(raw); gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        res = mc.process_page(img, template); sid = idr.read_id(gray)
        el = normalize_elective(res["subject"]); key = keys.get(el)
        nm = names_full.get((sid["ban"], sid["beon"])) or names_beon.get(sid["beon"], "")
        if key:
            g = grade_math(res["answers"], key)
            row = dict(page=i, name=nm, school=sid["school"], ban=sid["ban"], beon=sid["beon"],
                       elective=res["subject"], score=g["score"], max=key["total"],
                       wrong=g["wrong"], answers=res["answers"])
        else:
            row = dict(page=i, name=nm, school=sid["school"], ban=sid["ban"], beon=sid["beon"],
                       elective=res["subject"], score=None, max=100, wrong=[], answers=res["answers"])
        warns = []
        if "?" in f"{sid['school']}{sid['ban']}{sid['beon']}":
            warns.append("수험번호")
        if not key:
            warns.append("선택과목 미검출/키없음")
        nblank = sum(1 for q in range(1, 31) if row["answers"].get(q, 0) == 0)
        ndup = sum(1 for v in row["answers"].values() if v == -1)
        if nblank >= 8:
            warns.append(f"미마킹{nblank}")
        if ndup:
            warns.append(f"중복{ndup}")
        row["warn"] = " ".join(warns)
        rows.append(row)
        print(f"  p{i:>2}: {nm:<5} {sid['ban']}-{sid['beon']} {res['subject'] or '?':<7} "
              f"{row['score'] if row['score'] is not None else '-'}/{row['max']}")

    sc = [r["score"] for r in rows if r["score"] is not None]
    if sc:
        print(f"\n응시 {len(sc)}명  평균 {sum(sc)/len(sc):.1f}  최고 {max(sc)}  최저 {min(sc)}")
    xlsx, csvp = write_outputs(rows, args.out, args.pdf.stem)
    flagged = sorted({r["page"] for r in rows if r.get("warn")})
    if flagged:
        mc.oc.save_review_images(args.pdf, flagged, args.out, args.dpi)
        print(f"검토용 카드 이미지 {len(flagged)}장 저장 (review_imgs/)")
    print(f"완료:\n  {xlsx}\n  {csvp}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
