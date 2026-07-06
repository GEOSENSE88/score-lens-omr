# -*- coding: utf-8 -*-
"""
채점 결과 → 통합 성적 워크북 (score_lens / omr_scorer)

OMR 리딩·채점 결과를 학교 업무 양식에 맞춰 하나의 .xlsx 로 떨군다.
시트 구성:
  1) 대교협      — 대교협 모의고사 통합양식(가채점) [저장] 시트와 동일한 A:S 열 배치.
  2) UNIV        — 학년별 레이아웃(1·2학년=원/표/백/등×과목, 3학년=수능형 공통+선택).
  3) 정오표_<과목> — 과목별 문항 O/X 표(학생 선택답안 + 정오 색칠).

OMR 은 '원점수'만 채운다. 표준점수·백분위·등급은 공식 성적표 발표 후 입력하는
값이므로 공란으로 둔다(가채점 단계).

── 입력 레코드 스키마 ──────────────────────────────────────────────
student = {
  "학년": 3, "반": "01", "번호": "01", "이름": "홍길동", "계열번호": "",
  "한국사": {"원점수": 18},
  "국어":  {"선택": "화법과작문"|"국어", "공통원": 62, "선택원": 22, "원점수": 84,
             "정오": {q: {"답": 3, "정답": 3, "배점": 2, "ok": True}}},
  "수학":  {"선택": "확률과통계"|"수학", "공통원": .., "선택원": .., "원점수": .., "정오": {...}},
  "영어":  {"원점수": 74, "정오": {...}},
  "탐구":  [{"과목": "한국지리", "원점수": 29, "정오": {...}},
            {"과목": "생활과윤리", "원점수": 34, "정오": {...}}],
  "제2외국어": {"과목": "", "원점수": None},
}
학년만 필수. 없는 과목 키는 생략하면 해당 칸은 공란.
탐구는 고3=선택 2과목 리스트 / 고1·2=[통합사회, 통합과학] 로 채운다.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as gc

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEAD_FILL = PatternFill("solid", fgColor="305496")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
GRAY = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _grade_int(students: list[dict]) -> int:
    for s in students:
        try:
            return int(str(s.get("학년", "")).strip().split(".")[0])
        except (ValueError, TypeError):
            continue
    return 3


def _num(v):
    return v if isinstance(v, (int, float)) else (v or "")


def _tamgu(s: dict, i: int) -> dict:
    t = s.get("탐구") or []
    return t[i] if i < len(t) else {}


# ── 1) 대교협 [저장] 레이아웃 (A:S) ────────────────────────────────
DAEGYO_TOP = [
    ("A", "학년도"), ("B", "학년"), ("C", "반"), ("D", "번호"), ("E", "이름"),
    ("F", "한국사"), ("G", "국어"), ("J", "수학"), ("M", "영어"),
    ("N", "탐구영역1"), ("P", "탐구영역2"), ("R", "제2외국어"),
]
DAEGYO_SUB = {
    "F": "원점수", "G": "영역", "H": "공통\n원점수", "I": "선택\n원점수",
    "J": "영역", "K": "공통\n원점수", "L": "선택\n원점수", "M": "원점수",
    "N": "과목명", "O": "원점수", "P": "과목명", "Q": "원점수",
    "R": "과목명", "S": "원점수",
}


def _write_daegyo(ws, students: list[dict], exam_year):
    ws.merge_cells("A1:E1")
    ws["A1"] = "가채점"
    ws["A1"].font = Font(bold=True)
    for col, label in DAEGYO_TOP:
        c = ws[f"{col}2"]
        c.value = label
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = CENTER
    for merge in ("G2:I2", "J2:L2", "N2:O2", "P2:Q2", "R2:S2"):
        ws.merge_cells(merge)
    for col in "ABCDE":
        ws.merge_cells(f"{col}2:{col}3")
    for col, label in DAEGYO_SUB.items():
        c = ws[f"{col}3"]
        c.value = label
        c.font = Font(bold=True)
        c.fill = SUB_FILL
        c.alignment = CENTER
    ws["F2"].alignment = CENTER
    ws.merge_cells("F2:F2")

    r = 4
    for s in students:
        kor = s.get("국어", {})
        mat = s.get("수학", {})
        t1, t2 = _tamgu(s, 0), _tamgu(s, 1)
        l2 = s.get("제2외국어", {})
        vals = {
            "A": exam_year, "B": s.get("학년", ""), "C": s.get("반", ""),
            "D": s.get("번호", ""), "E": s.get("이름", ""),
            "F": _num(s.get("한국사", {}).get("원점수")),
            "G": kor.get("선택", "") or ("국어" if kor else ""),
            "H": _num(kor.get("공통원")), "I": _num(kor.get("선택원")),
            "J": mat.get("선택", "") or ("수학" if mat else ""),
            "K": _num(mat.get("공통원")), "L": _num(mat.get("선택원")),
            "M": _num(s.get("영어", {}).get("원점수")),
            "N": t1.get("과목", ""), "O": _num(t1.get("원점수")),
            "P": t2.get("과목", ""), "Q": _num(t2.get("원점수")),
            "R": l2.get("과목", ""), "S": _num(l2.get("원점수")),
        }
        for col, v in vals.items():
            cell = ws[f"{col}{r}"]
            cell.value = v
            cell.alignment = CENTER
        r += 1
    widths = {"A": 8, "B": 5, "C": 5, "D": 5, "E": 9, "G": 11, "J": 11,
              "N": 11, "P": 11, "R": 11}
    for col in "ABCDEFGHIJKLMNOPQRS":
        ws.column_dimensions[col].width = widths.get(col, 8)
    ws.freeze_panes = "A4"


# ── 2) UNIV 학년별 레이아웃 ────────────────────────────────────────
# 각 (학년) → 그룹 리스트. 그룹 = (헤더라벨, 시작열idx, 컬럼정의)
# 컬럼정의: 순서대로 붙는 하위헤더 라벨. 값은 채우기 함수가 결정.
def _univ_layout(grade: int):
    """returns (groups, fill_fn). groups=[(라벨, [하위헤더...])], A~E 공통 선행."""
    if grade in (1,):
        groups = [
            ("국어", ["원점수", "표준점수", "백분위", "등급"]),
            ("수학", ["원점수", "표준점수", "백분위", "등급"]),
            ("영어", ["원점수", "표준점수", "백분위", "등급"]),
            ("사회탐구", ["원점수", "표준점수", "백분위", "등급"]),
            ("과학탐구", ["원점수", "표준점수", "백분위", "등급"]),
            ("한국사", ["원점수", "표준점수", "백분위", "등급"]),
        ]
    elif grade == 2:
        groups = [
            ("국어", ["원점수", "표준점수", "백분위", "등급"]),
            ("수학", ["원점수", "표준점수", "백분위", "등급"]),
            ("영어", ["원점수", "표준점수", "백분위", "등급"]),
            ("사회탐구", ["원점수", "표준점수", "백분위", "등급"]),
            ("과학탐구", ["원점수", "표준점수", "백분위", "등급"]),
            ("직업탐구", ["원점수", "표준점수", "백분위", "등급"]),
            ("한국사", ["원점수", "표준점수", "백분위", "등급"]),
            ("제2외국어", ["과목명", "원점수", "표준점수", "백분위", "등급"]),
        ]
    else:  # 3학년 = 수능형
        groups = [
            ("국어", ["국어선택", "공통원점수", "선택원점수", "원점수총점", "표준점수", "백분위", "등급"]),
            ("수학", ["수학선택", "공통원점수", "선택원점수", "원점수총점", "표준점수", "백분위", "등급"]),
            ("영어", ["원점수", "표준점수", "백분위", "등급"]),
            ("선택1", ["과목명", "원점수", "표준점수", "백분위", "등급"]),
            ("선택2", ["과목명", "원점수", "표준점수", "백분위", "등급"]),
            ("한국사", ["원점수", "표준점수", "백분위", "등급"]),
            ("제2외국어", ["과목명", "원점수", "표준점수", "백분위", "등급"]),
        ]
    return groups


def _subject_data(s: dict, group_label: str):
    """UNIV 그룹 라벨 → 해당 과목 데이터 dict (없으면 None)."""
    tam = s.get("탐구") or []

    def find_tamgu(names):
        for t in tam:
            if t.get("과목") in names or t.get("영역") in names:
                return t
        return None

    if group_label in ("국어", "수학", "영어", "한국사"):
        return s.get(group_label)
    # ⚠️ 위치 폴백(tam[0]/tam[1]) 금지 — 통합과학만 응시한 학생이
    #    사회탐구 열에 잘못 표기되는 사고를 막는다. 과목명으로만 매칭.
    if group_label == "사회탐구":
        return find_tamgu({"통합사회", "사회탐구"})
    if group_label == "과학탐구":
        return find_tamgu({"통합과학", "과학탐구"})
    if group_label in ("선택1", "선택2"):
        idx = 0 if group_label == "선택1" else 1
        return tam[idx] if idx < len(tam) else None
    if group_label == "제2외국어":
        return s.get("제2외국어")
    return None


def _univ_values(s: dict, grade: int, group_label: str, sub: str):
    """한 학생의 (그룹, 하위헤더) → 채울 값.
    표준점수/백분위/등급은 EBSi 등급컷 기반 '예상치'가 있으면 채운다."""
    if sub in ("표준점수", "백분위", "등급"):
        d = _subject_data(s, group_label)
        return d.get(f"예상{sub}", "") if d else ""
    kor, mat = s.get("국어", {}), s.get("수학", {})
    if group_label == "국어":
        if sub in ("원점수",):
            return _num(kor.get("원점수"))
        if sub == "국어선택":
            return kor.get("선택", "") or "국어"
        if sub == "공통원점수":
            return _num(kor.get("공통원"))
        if sub == "선택원점수":
            return _num(kor.get("선택원"))
        if sub == "원점수총점":
            return _num(kor.get("원점수"))
    if group_label == "수학":
        if sub == "원점수":
            return _num(mat.get("원점수"))
        if sub == "수학선택":
            return mat.get("선택", "") or "수학"
        if sub == "공통원점수":
            return _num(mat.get("공통원"))
        if sub == "선택원점수":
            return _num(mat.get("선택원"))
        if sub == "원점수총점":
            return _num(mat.get("원점수"))
    if group_label == "영어" and sub == "원점수":
        return _num(s.get("영어", {}).get("원점수"))
    if group_label == "한국사" and sub == "원점수":
        return _num(s.get("한국사", {}).get("원점수"))
    # 탐구 매핑: 고1·2 = 사회탐구/과학탐구(통합), 고3 = 선택1/선택2
    tam = s.get("탐구") or []
    def find_tamgu(names):
        for t in tam:
            if t.get("과목") in names or t.get("영역") in names:
                return t
        return None
    if group_label == "사회탐구":
        t = find_tamgu({"통합사회", "사회탐구"})     # 위치 폴백 금지 (과목명 매칭만)
        if t:
            if sub == "원점수":
                return _num(t.get("원점수"))
    if group_label == "과학탐구":
        t = find_tamgu({"통합과학", "과학탐구"})     # 위치 폴백 금지
        if t and sub == "원점수":
            return _num(t.get("원점수"))
    if group_label in ("선택1", "선택2"):
        idx = 0 if group_label == "선택1" else 1
        t = tam[idx] if idx < len(tam) else {}
        if sub == "과목명":
            return t.get("과목", "")
        if sub == "원점수":
            return _num(t.get("원점수"))
    if group_label == "제2외국어":
        l2 = s.get("제2외국어", {})
        if sub == "과목명":
            return l2.get("과목", "")
        if sub == "원점수":
            return _num(l2.get("원점수"))
    if group_label == "직업탐구" and sub == "원점수":
        return ""
    return ""


def _write_univ(ws, students: list[dict], grade: int):
    base = ["계열번호", "학년", "반", "번호", "이름"]
    for i, label in enumerate(base, 1):
        c = ws.cell(1, i, label)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD_FILL; c.alignment = CENTER
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
    groups = _univ_layout(grade)
    col = 6
    col_index = []  # (group_label, sub, col)
    for label, subs in groups:
        start = col
        gc0 = ws.cell(1, col, label)
        gc0.font = Font(bold=True, color="FFFFFF"); gc0.fill = HEAD_FILL; gc0.alignment = CENTER
        for sub in subs:
            sc = ws.cell(2, col, sub)
            sc.font = Font(bold=True); sc.fill = SUB_FILL; sc.alignment = CENTER
            col_index.append((label, sub, col))
            col += 1
        if col - 1 > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col - 1)

    r = 3
    for s in students:
        ws.cell(r, 1, s.get("계열번호", "")).alignment = CENTER
        ws.cell(r, 2, s.get("학년", "")).alignment = CENTER
        ws.cell(r, 3, s.get("반", "")).alignment = CENTER
        ws.cell(r, 4, s.get("번호", "")).alignment = CENTER
        ws.cell(r, 5, s.get("이름", "")).alignment = CENTER
        for label, sub, c in col_index:
            cell = ws.cell(r, c, _univ_values(s, grade, label, sub))
            cell.alignment = CENTER
        r += 1
    for c in range(1, col):
        ws.column_dimensions[gc(c)].width = 9 if c > 5 else 7
    ws.freeze_panes = "F3"


# ── 3) 과목별 정오표 ──────────────────────────────────────────────
def _write_marksheet(ws, subject: str, rows: list[dict], nq: int):
    NB = 6                                  # 문항 앞 기본 컬럼 수
    # 1행: 범례 (셀 숫자 = 학생이 마킹한 답안)
    ws.cell(1, 1, "숫자 = 학생이 마킹한 답 ·  초록 = 정답  빨강 = 오답  "
                  "· = 미마킹  중 = 중복마킹  회색 = 정답키 없음").font = Font(bold=True, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NB + nq)

    headers = ["반", "번호", "이름", "원점수", "만점", "틀린문항"] + [str(q) for q in range(1, nq + 1)]
    for i, h in enumerate(headers, 1):
        c = ws.cell(2, i, h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD_FILL; c.alignment = CENTER
    for ri, row in enumerate(rows, 3):
        jeongo = row.get("정오", {})
        wrong = [q for q, d in jeongo.items() if not d.get("ok") and d.get("답") not in (0, -1, None, "")]
        base = [row.get("반", ""), row.get("번호", ""), row.get("이름", ""),
                _num(row.get("원점수")), _num(row.get("만점")), " ".join(map(str, sorted(wrong)))]
        for ci, v in enumerate(base, 1):
            ws.cell(ri, ci, v).alignment = CENTER
        for q in range(1, nq + 1):
            d = jeongo.get(q) or jeongo.get(str(q))
            cell = ws.cell(ri, NB + q)
            cell.alignment = CENTER
            if not d:
                cell.fill = GRAY
                continue
            stu, ok = d.get("답"), d.get("ok")
            if stu in (0, "", None):
                cell.value = "·"            # 미마킹
            elif stu == -1:
                cell.value = "중"           # 중복마킹
            else:
                cell.value = stu            # 학생이 마킹한 답 (정답/오답 모두 색으로 구분)
            cell.fill = GREEN if ok else RED
    ws.freeze_panes = ws.cell(3, NB + 1).coordinate
    ws.row_dimensions[1].height = 15
    for c in range(1, NB + nq + 1):
        ws.column_dimensions[gc(c)].width = 5.4 if c > NB else (9 if c == 3 else 6)


def _write_cuts_sheet(ws, cuts_by_subject: dict):
    """예상등급컷 근거 시트 (EBSi 출처 표기)."""
    ws.append(["※ EBSi 예상/확정 등급컷 기반 추정치입니다. 공식 성적표를 대체하지 않습니다."])
    ws["A1"].font = Font(bold=True, color="CB3534")
    r = 3
    for subject, cut in cuts_by_subject.items():
        ws.cell(r, 1, subject).font = Font(bold=True, size=12)
        meta = []
        if cut.get("avg") is not None:
            meta.append(f"평균 {cut['avg']}")
        if cut.get("std") is not None:
            meta.append(f"표준편차 {cut['std']}")
        ws.cell(r, 3, " / ".join(meta))
        r += 1
        headers = ["등급", "원점수컷", "표준점수", "백분위"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(r, ci, h)
            c.font = Font(bold=True)
            c.fill = SUB_FILL
            c.alignment = CENTER
        r += 1
        for row in cut.get("cuts", []):
            label = "최고점" if row.get("grade") == 0 else row.get("grade")
            vals = [label, row.get("raw"), row.get("std_score"), row.get("pct")]
            for ci, v in enumerate(vals, 1):
                ws.cell(r, ci, "-" if v is None else v).alignment = CENTER
            r += 1
        r += 1
    for col, w in zip("ABCD", [8, 10, 10, 8]):
        ws.column_dimensions[col].width = w


def build_workbook(students: list[dict], out_path: str | Path,
                   exam_year: int | str | None = None,
                   marksheets: dict[str, tuple[list[dict], int]] | None = None,
                   cuts_by_subject: dict | None = None) -> Path:
    """통합 워크북 생성.

    students   : 위 스키마의 학생 레코드 리스트
    marksheets : {과목: (rows, 문항수)} — 정오표용. rows 는 {반,번호,이름,원점수,만점,정오}.
                 생략 시 students 의 과목별 '정오'에서 자동 구성.
    cuts_by_subject : EBSi 등급컷(과목별) — 있으면 '예상등급컷' 근거 시트 추가.
    """
    grade = _grade_int(students)
    if exam_year is None:
        exam_year = {1: 2029, 2: 2028, 3: 2027}.get(grade, 2027)  # 수능 응시연도 기준

    wb = Workbook()
    _write_daegyo(wb.active, students, exam_year)
    wb.active.title = "대교협"
    _write_univ(wb.create_sheet(f"UNIV {grade}학년"), students, grade)
    if cuts_by_subject:
        _write_cuts_sheet(wb.create_sheet("예상등급컷"), cuts_by_subject)

    # 정오표: 명시 marksheets 우선, 없으면 students 에서 과목별로 모은다
    if marksheets is None:
        marksheets = {}
        subj_keys = ["국어", "수학", "영어", "한국사"]
        collected: dict[str, list[dict]] = {k: [] for k in subj_keys}
        maxq: dict[str, int] = {k: 0 for k in subj_keys}
        tamgu_rows: dict[str, list[dict]] = {}
        tamgu_nq: dict[str, int] = {}
        for s in students:
            ident = {"반": s.get("반", ""), "번호": s.get("번호", ""), "이름": s.get("이름", "")}
            for k in subj_keys:
                sub = s.get(k)
                if sub and sub.get("정오"):
                    jeongo = {int(q): v for q, v in sub["정오"].items()}
                    collected[k].append({**ident, "원점수": sub.get("원점수"),
                                         "만점": sub.get("만점"), "정오": jeongo})
                    maxq[k] = max(maxq[k], max(jeongo) if jeongo else 0)
            for t in (s.get("탐구") or []):
                if t.get("정오"):
                    nm = t.get("과목", "탐구")
                    jeongo = {int(q): v for q, v in t["정오"].items()}
                    tamgu_rows.setdefault(nm, []).append({**ident, "원점수": t.get("원점수"),
                                                          "만점": t.get("만점"), "정오": jeongo})
                    tamgu_nq[nm] = max(tamgu_nq.get(nm, 0), max(jeongo) if jeongo else 0)
        for k in subj_keys:
            if collected[k]:
                marksheets[k] = (collected[k], maxq[k])
        for nm, rows in tamgu_rows.items():
            marksheets[nm] = (rows, tamgu_nq[nm])

    for subject, (rows, nq) in marksheets.items():
        if not rows or nq <= 0:
            continue
        title = f"정오표_{subject}"[:31]
        _write_marksheet(wb.create_sheet(title), subject, rows, nq)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
