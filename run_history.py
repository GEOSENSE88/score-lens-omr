# -*- coding: utf-8 -*-
"""Korean History OMR pipeline: PDF/images -> 20 answers -> scoring."""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

import cv2

import history_core as hc
import id_reader as idr
from run_math import load_name_map


def load_history_key(keys_dir: Path, exam_id: str | None) -> dict | None:
    candidates = []
    for p in keys_dir.glob("*.json"):
        k = json.loads(p.read_text(encoding="utf-8"))
        if exam_id and str(k.get("exam_id") or k.get("irecord")) != str(exam_id):
            continue
        if k.get("subject") != "한국사":
            continue
        candidates.append(k)
    if not candidates:
        return None
    candidates.sort(key=lambda k: (not bool(k.get("complete")), -int(k.get("item_count", 0))))
    return candidates[0]


def grade_known(answers: dict[int, int], key: dict | None) -> dict:
    if not key:
        return dict(score=None, max_score=0, correct=[], wrong=[], blank=[], dup=[], per_q={})
    kans = {int(q): int(v) for q, v in key["answers"].items()}
    kpts = {int(q): int(v) for q, v in key["points"].items()}
    score, correct, wrong, blank, dup, per_q = 0, [], [], [], [], {}
    for q in sorted(kans):
        stu = answers.get(q, 0)
        ok = stu == kans[q]
        if stu == 0:
            blank.append(q)
        elif stu == -1:
            dup.append(q)
        elif ok:
            score += kpts[q]
            correct.append(q)
        else:
            wrong.append(q)
        per_q[q] = {"student": stu, "answer": kans[q], "points": kpts[q], "ok": ok}
    return dict(score=score, max_score=sum(kpts.values()), correct=correct, wrong=wrong, blank=blank, dup=dup, per_q=per_q)


def write_outputs(rows: list[dict], out: Path, stem: str, key: dict | None) -> tuple[Path, Path]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    out.mkdir(parents=True, exist_ok=True)
    headers = ["페이지", "학교번호", "반", "번호", "성명", "점수", "만점", "틀린문항"] + [str(q) for q in range(1, 21)] + ["확인필요"]

    csv_path = out / f"{stem}_한국사_판독표.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow(
                [r["page"], r["school"], r["ban"], r["beon"], r["name"], r["score"], r["max_score"], " ".join(map(str, r["wrong"]))]
                + [r["answers"].get(q, "") for q in range(1, 21)] + [r.get("warn", "")]
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "한국사 판독"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    gray = PatternFill("solid", fgColor="D9D9D9")
    known = set(int(q) for q in (key or {}).get("answers", {}))
    center = Alignment(horizontal="center")
    for row_idx, r in enumerate(rows, 2):
        base = [r["page"], r["school"], r["ban"], r["beon"], r["name"], r["score"], r["max_score"], " ".join(map(str, r["wrong"]))]
        for col_idx, value in enumerate(base, 1):
            ws.cell(row_idx, col_idx, value).alignment = center
        for q in range(1, 21):
            cell = ws.cell(row_idx, 8 + q, r["answers"].get(q, ""))
            cell.alignment = center
            if q in known:
                cell.fill = green if q not in r["wrong"] and q not in r["blank"] and q not in r["dup"] else red
            else:
                cell.fill = gray
    ws.freeze_panes = "I2"
    for col in range(1, 29):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 4.5 if col > 8 else 10
    xlsx_path = out / f"{stem}_한국사_판독표.xlsx"
    wb.save(xlsx_path)
    return xlsx_path, csv_path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("pdf", type=Path)
    ap.add_argument("--irecord", default=None)
    ap.add_argument("--keys-dir", type=Path, default=Path("keys"))
    ap.add_argument("--names", default=None, help="Existing score CSV for name join")
    ap.add_argument("--template", type=Path, default=None)
    ap.add_argument("--dpi", type=int, default=300)
    ap.add_argument("--out", type=Path, default=Path("output"))
    args = ap.parse_args()

    template = hc.load_template(args.template)
    key = load_history_key(args.keys_dir, args.irecord)
    if key:
        print(f"한국사 키: {key.get('item_count', len(key.get('answers', {})))}문항, complete={key.get('complete', False)}")
    else:
        print("한국사 키 없음: 판독표만 생성")
    names_full, names_beon = load_name_map(args.names)

    rows = []
    for pi, raw in hc.oc.iter_pdf_pages(args.pdf, args.dpi):
        i = pi + 1
        img = hc.oc.load_page(raw)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        sid = idr.read_id(gray)
        res = hc.process_page(img, template)
        g = grade_known(res["answers"], key)
        name = names_full.get((sid["ban"], sid["beon"])) or names_beon.get(sid["beon"], "")
        row = dict(page=i, school=sid["school"], ban=sid["ban"], beon=sid["beon"], name=name, answers=res["answers"], **g)
        warns = []
        if "?" in f"{sid['school']}{sid['ban']}{sid['beon']}":
            warns.append("수험번호")
        if len(g["blank"]) >= 5:
            warns.append(f"미마킹{len(g['blank'])}")
        if g["dup"]:
            warns.append(f"중복{len(g['dup'])}")
        row["warn"] = " ".join(warns)
        rows.append(row)
        print(f"p{i:02d} {name:<5} {sid['ban']}-{sid['beon']} score {row['score']}/{row['max_score']}")

    xlsx, csvp = write_outputs(rows, args.out, args.pdf.stem, key)
    flagged = sorted({r["page"] for r in rows if r.get("warn")})
    if flagged:
        hc.oc.save_review_images(args.pdf, flagged, args.out, args.dpi)
        print(f"검토용 카드 이미지 {len(flagged)}장 저장 (review_imgs/)")
    print(f"완료:\n  {xlsx}\n  {csvp}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
